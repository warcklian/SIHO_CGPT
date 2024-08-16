# cspell: disable
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import datetime
import time
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Solicitar el archivo Excel al usuario
Tk().withdraw()  # Ocultar la ventana principal de Tkinter
excel_file_path = askopenfilename(title="Seleccione el archivo Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])

if not excel_file_path:
    raise ValueError("No se seleccionó ningún archivo Excel. El programa no puede continuar.")

# Configurar las opciones de Chrome para conectar a la instancia existente
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

# Ruta al ejecutable de ChromeDriver
service = Service(executable_path="./chromedriver.exe")
driver = webdriver.Chrome(service=service, options=chrome_options)

# Obtener el identificador de la ventana actual (suponiendo que el formulario ya está abierto en esta ventana)
current_window_handle = driver.current_window_handle

# Leer el archivo Excel
df = pd.read_excel(excel_file_path, engine='openpyxl')

# Crear una instancia de WebDriverWait
wait = WebDriverWait(driver, 10)

# Abrir el archivo Excel para escribir las fechas
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Buscar la columna que contiene el encabezado 'Date'
date_column_letter = None
for column in sheet.iter_cols(1, sheet.max_column):
    if column[0].value == 'Date':
        date_column_letter = column[0].column_letter
        break

if not date_column_letter:
    raise ValueError("No se encontró la columna 'Date' en el archivo Excel.")

# Iterar sobre cada fila del DataFrame y completar el formulario
for index, row in df.iterrows():
    # Cambiar a la pestaña o ventana activa
    driver.switch_to.window(current_window_handle)

    # Esperar a que los campos del formulario estén presentes
    wait.until(EC.presence_of_element_located((By.ID, "ur")))

    # Completar los campos del formulario con los datos del archivo Excel
    driver.find_element(By.ID, "ur").clear()
    driver.find_element(By.ID, "ur").send_keys("600 -->SUBSECRETARÍA DE EDUCACIÓN MEDIA SUPERIOR")  # Valor fijo
    
    driver.find_element(By.ID, "programa").clear()
    driver.find_element(By.ID, "programa").send_keys(row['Programa'])
    
    driver.find_element(By.ID, "oficio").clear()
    driver.find_element(By.ID, "oficio").send_keys(row['Oficio de afectación presupuestal'])
    
    driver.find_element(By.ID, "curp").clear()
    driver.find_element(By.ID, "curp").send_keys(row['CURP'])
    
    driver.find_element(By.ID, "rfc").clear()
    driver.find_element(By.ID, "rfc").send_keys(row['RFC y Homoclave'])
    
    driver.find_element(By.ID, "primerApellido").clear()
    driver.find_element(By.ID, "primerApellido").send_keys(row['Primer apellido'])
    
    driver.find_element(By.ID, "nombre").clear()
    driver.find_element(By.ID, "nombre").send_keys(row['Nombre(s)'])
    
    driver.find_element(By.ID, "nacionalidad").clear()
    driver.find_element(By.ID, "nacionalidad").send_keys(row['Nacionalidad'])
    
    driver.find_element(By.ID, "sexo").clear()
    driver.find_element(By.ID, "sexo").send_keys(row['Sexo'])
    
    driver.find_element(By.ID, "escolaridad").clear()
    driver.find_element(By.ID, "escolaridad").send_keys(row['Escolaridad'])
    
    driver.find_element(By.ID, "estadoNacimiento").clear()
    driver.find_element(By.ID, "estadoNacimiento").send_keys(row['Estado de Nacimiento'])
    
    driver.find_element(By.ID, "especifique").clear()
    driver.find_element(By.ID, "especifique").send_keys(str(row['Especifique']))
    
    driver.find_element(By.ID, "banco").clear()
    driver.find_element(By.ID, "banco").send_keys(row['Banco o CLABE'])
    
    driver.find_element(By.ID, "clabe").clear()
    driver.find_element(By.ID, "clabe").send_keys(row['CLABE'])
    
    driver.find_element(By.ID, "fechaInicio").clear()
    driver.find_element(By.ID, "fechaInicio").send_keys(row['Fecha de inicio del contrato'])
    
    driver.find_element(By.ID, "fechaTermino").clear()
    driver.find_element(By.ID, "fechaTermino").send_keys(row['Fecha de término del contrato'])
    
    driver.find_element(By.ID, "nivelContrato").clear()
    driver.find_element(By.ID, "nivelContrato").send_keys(row['Nivel del Contrato'])
    
    driver.find_element(By.ID, "montoBruto").clear()
    driver.find_element(By.ID, "montoBruto").send_keys(str(row['Monto Bruto Mensual']))
    
    driver.find_element(By.ID, "fina").clear()
    driver.find_element(By.ID, "fina").send_keys(row['Fina'])
    
    # Marcar la casilla si 'Compatibilidad' es True
    compatibilidad_checkbox = driver.find_element(By.ID, "compatibilidad")
    if row['Compatibilidad']:
        if not compatibilidad_checkbox.is_selected():
            compatibilidad_checkbox.click()
    else:
        if compatibilidad_checkbox.is_selected():
            compatibilidad_checkbox.click()
    
    driver.find_element(By.ID, "infoSoporte").clear()
    driver.find_element(By.ID, "infoSoporte").send_keys(row['Conocimientos requeridos por la dependencia'])
    
    # Esperar un poco para evitar enviar datos demasiado rápido
    time.sleep(1)
    
    # Enviar el formulario
    driver.find_element(By.XPATH, "//button[@class='accept-btn']").click()
    
    # Registrar la fecha y hora actual en la columna 'Date' correspondiente
    current_datetime = datetime.datetime.now()
    sheet[f'{date_column_letter}{index + 2}'] = current_datetime

# Guardar el archivo Excel con las fechas y horas agregadas
workbook.save(excel_file_path)

# No cerrar el navegador
# driver.quit()
