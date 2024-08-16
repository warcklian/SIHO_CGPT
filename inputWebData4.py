import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import datetime
import time
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Solicitar el archivo Excel al usuario
Tk().withdraw()  # Ocultar la ventana principal de Tkinter
excel_file_path = askopenfilename(title="Seleccione el archivo Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])

if not excel_file_path:
    raise ValueError("No se seleccionó ningún archivo Excel. El programa no puede continuar.")

# Configurar las opciones de Chrome para conectar a la instancia existente
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

# Ruta al ejecutable de ChromeDriver
service = Service(executable_path="./chromedriver.exe")
driver = webdriver.Chrome(service=service, options=chrome_options)

# Obtener el identificador de la ventana actual (suponiendo que el formulario ya está abierto en esta ventana)
current_window_handle = driver.current_window_handle

# Leer el archivo Excel
df = pd.read_excel(excel_file_path, engine='openpyxl')

# Crear una instancia de WebDriverWait
wait = WebDriverWait(driver, 10)

# Abrir el archivo Excel para escribir las fechas
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Buscar la columna que contiene el encabezado 'Date'
date_column_letter = None
for column in sheet.iter_cols(1, sheet.max_column):
    if column[0].value == 'Date':
        date_column_letter = column[0].column_letter
        break

if not date_column_letter:
    raise ValueError("No se encontró la columna 'Date' en el archivo Excel.")

# Iterar sobre cada fila del DataFrame y completar el formulario
for index, row in df.iterrows():
    # Cambiar a la pestaña o ventana activa
    driver.switch_to.window(current_window_handle)

    # Esperar a que los campos del formulario estén presentes
    wait.until(EC.presence_of_element_located((By.NAME, "first_name")))

    # Completar los campos del formulario con los datos del archivo Excel
    driver.find_element(By.NAME, "first_name").clear()
    driver.find_element(By.NAME, "first_name").send_keys(row['First Name'])
    
    driver.find_element(By.NAME, "last_name_paternal").clear()
    driver.find_element(By.NAME, "last_name_paternal").send_keys(row['Last Name (Paternal)'])
    
    driver.find_element(By.NAME, "last_name_maternal").clear()
    driver.find_element(By.NAME, "last_name_maternal").send_keys(row['Last Name (Maternal)'])
    
    driver.find_element(By.NAME, "rfc").clear()
    driver.find_element(By.NAME, "rfc").send_keys(row['RFC'])
    
    driver.find_element(By.NAME, "address").clear()
    driver.find_element(By.NAME, "address").send_keys(row['Address'])
    
    driver.find_element(By.NAME, "postal_code").clear()
    driver.find_element(By.NAME, "postal_code").send_keys(str(row['Postal Code']))  # Convertir a cadena
    
    # Verificar si el código postal fue enviado correctamente antes de continuar
    wait.until(EC.text_to_be_present_in_element_value((By.NAME, "postal_code"), str(row['Postal Code'])))

    driver.find_element(By.NAME, "phone").clear()
    driver.find_element(By.NAME, "phone").send_keys(row['Phone'])
    
    driver.find_element(By.NAME, "email").clear()
    driver.find_element(By.NAME, "email").send_keys(row['Email'])
    
    # Esperar un poco para evitar enviar datos demasiado rápido
    time.sleep(1)
    
    # Enviar el formulario
    driver.find_element(By.NAME, "submit_button").click()
    
    # Registrar la fecha y hora actual en la columna 'Date' correspondiente
    current_datetime = datetime.datetime.now()
    sheet[f'{date_column_letter}{index + 2}'] = current_datetime

# Guardar el archivo Excel con las fechas y horas agregadas
workbook.save(excel_file_path)

# No cerrar el navegador
# driver.quit()
