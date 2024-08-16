import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading

class FormFillerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Formulario de Relleno")
        self.root.geometry("500x450")

        self.is_running = False  # Variable para controlar la ejecución

        # Descripción del programa
        self.description_text = (
            "Este programa permite automatizar el llenado de formularios web utilizando datos de un archivo Excel.\n\n"
            "Funcionalidades:\n"
            "1. Selecciona un archivo Excel que contiene los datos a ingresar.\n"
            "2. El programa abrirá un navegador web y completará el formulario con los datos del archivo.\n"
            "3. Después de enviar cada formulario, registrará la fecha, hora y un número de trámite en el archivo Excel.\n\n"
            "Estructura del archivo Excel:\n"
            "El archivo debe contener las siguientes columnas:\n"
            "- First Name\n"
            "- Last Name (Paternal)\n"
            "- Last Name (Maternal)\n"
            "- RFC\n"
            "- Address\n"
            "- Postal Code\n"
            "- Phone\n"
            "- Email\n"
            "- Start\n"
            "- End\n"
            "\nAdemás, debe haber una columna llamada 'Date' en la que se registrarán las fechas de envío,\n"
            "una columna llamada 'Time' para registrar la hora, minutos y segundos del envío, y\n"
            "una columna llamada 'Tramite' para registrar el número de trámite."
        )

        self.description_label = tk.Label(root, text="Descripción del Programa", font=("Arial", 12, "bold"))
        self.description_label.pack(pady=10)

        self.description_text_widget = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=10, width=60)
        self.description_text_widget.pack(pady=10)
        self.description_text_widget.insert(tk.INSERT, self.description_text)
        self.description_text_widget.configure(state=tk.DISABLED)  # Solo lectura

        self.select_button = tk.Button(root, text="Seleccionar archivo", command=self.select_file)
        self.select_button.pack(pady=10)

        self.process_button = tk.Button(root, text="Procesar archivo", command=self.start_process, state=tk.DISABLED)
        self.process_button.pack(pady=10)

        self.stop_button = tk.Button(root, text="Detener", command=self.stop_process, state=tk.DISABLED)
        self.stop_button.pack(pady=10)

        self.excel_file_path = None

    def select_file(self):
        self.excel_file_path = filedialog.askopenfilename(
            title="Seleccione el archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx;*.xls")]
        )
        if self.excel_file_path:
            self.process_button.config(state=tk.NORMAL)

    def start_process(self):
        self.is_running = True
        self.process_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        threading.Thread(target=self.process_file).start()

    def stop_process(self):
        self.is_running = False
        self.stop_button.config(state=tk.DISABLED)

    def process_file(self):
        if not self.excel_file_path:
            messagebox.showerror("Error", "No se ha seleccionado ningún archivo Excel.")
            return

        try:
            # Configurar las opciones de Chrome para conectar a la instancia existente
            chrome_options = Options()
            chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

            # Ruta al ejecutable de ChromeDriver
            service = Service(executable_path="./chromedriver.exe")
            driver = webdriver.Chrome(service=service, options=chrome_options)

            # Obtener el identificador de la ventana actual (suponiendo que el formulario ya está abierto en esta ventana)
            current_window_handle = driver.current_window_handle

            # Leer el archivo Excel
            df = pd.read_excel(self.excel_file_path, engine='openpyxl')

            # Crear una instancia de WebDriverWait
            wait = WebDriverWait(driver, 10)

            # Abrir el archivo Excel para escribir las fechas, horas y números de trámite
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sheet = workbook.active

            # Buscar las columnas 'Date', 'Time', y 'Tramite'
            date_column_letter = None
            time_column_letter = None
            tramite_column_letter = None
            for column in sheet.iter_cols(1, sheet.max_column):
                if column[0].value == 'Date':
                    date_column_letter = column[0].column_letter
                elif column[0].value == 'Time':
                    time_column_letter = column[0].column_letter
                elif column[0].value == 'Tramite':
                    tramite_column_letter = column[0].column_letter

            if not date_column_letter:
                raise ValueError("No se encontró la columna 'Date' en el archivo Excel.")
            if not time_column_letter:
                raise ValueError("No se encontró la columna 'Time' en el archivo Excel.")
            if not tramite_column_letter:
                raise ValueError("No se encontró la columna 'Tramite' en el archivo Excel.")

            # Iterar sobre cada fila del DataFrame y completar el formulario
            for index, row in df.iterrows():
                if not self.is_running:
                    messagebox.showinfo("Proceso Detenido", "El proceso ha sido detenido por el usuario.")
                    break  # Detener la ejecución si se ha presionado el botón "Detener"
                
                # Cambiar a la pestaña o ventana activa
                driver.switch_to.window(current_window_handle)

                # Esperar a que los campos del formulario estén presentes
                wait.until(EC.presence_of_element_located((By.NAME, "first_name")))

                # Completar los campos del formulario con los datos del archivo Excel
                if self.is_field_available(driver, "first_name"):
                    driver.find_element(By.NAME, "first_name").clear()
                    driver.find_element(By.NAME, "first_name").send_keys(row['First Name'])
                
                if self.is_field_available(driver, "last_name_paternal"):
                    driver.find_element(By.NAME, "last_name_paternal").clear()
                    driver.find_element(By.NAME, "last_name_paternal").send_keys(row['Last Name (Paternal)'])
                
                if self.is_field_available(driver, "last_name_maternal"):
                    driver.find_element(By.NAME, "last_name_maternal").clear()
                    driver.find_element(By.NAME, "last_name_maternal").send_keys(row['Last Name (Maternal)'])
                
                if self.is_field_available(driver, "rfc"):
                    driver.find_element(By.NAME, "rfc").clear()
                    driver.find_element(By.NAME, "rfc").send_keys(row['RFC'])
                
                if self.is_field_available(driver, "address"):
                    driver.find_element(By.NAME, "address").clear()
                    driver.find_element(By.NAME, "address").send_keys(row['Address'])
                
                if self.is_field_available(driver, "postal_code"):
                    driver.find_element(By.NAME, "postal_code").clear()
                    driver.find_element(By.NAME, "postal_code").send_keys(str(row['Postal Code']))  # Convertir a cadena
                
                    # Verificar si el código postal fue enviado correctamente antes de continuar
                    wait.until(EC.text_to_be_present_in_element_value((By.NAME, "postal_code"), str(row['Postal Code'])))

                if self.is_field_available(driver, "phone"):
                    driver.find_element(By.NAME, "phone").clear()
                    driver.find_element(By.NAME, "phone").send_keys(row['Phone'])
                
                if self.is_field_available(driver, "email"):
                    driver.find_element(By.NAME, "email").clear()
                    driver.find_element(By.NAME, "email").send_keys(row['Email'])
                
                # Convertir y formatear las fechas 'Start' y 'End' en formato 'dd/mm/yyyy'
                start_date = row['Start'].strftime("%d/%m/%Y")
                end_date = row['End'].strftime("%d/%m/%Y")

                # Completar los campos de fechas 'Start' y 'End'
                if self.is_field_available(driver, "start"):
                    driver.find_element(By.NAME, "start").clear()
                    driver.find_element(By.NAME, "start").send_keys(start_date)

                if self.is_field_available(driver, "end"):
                    driver.find_element(By.NAME, "end").clear()
                    driver.find_element(By.NAME, "end").send_keys(end_date)

                # Obtener el número de trámite generado por el HTML
                if self.is_field_available(driver, "tramite_number"):
                    tramite_number = driver.find_element(By.NAME, "tramite_number").get_attribute("value")
                
                    # Registrar el número de trámite en la columna correspondiente del Excel
                    sheet[f'{tramite_column_letter}{index + 2}'] = tramite_number
                
                # Obtener la fecha y hora actuales
                now = datetime.datetime.now()
                current_date = now.date()
                current_time = now.strftime("%H:%M:%S")
                
                # Registrar la fecha y hora en las columnas correspondientes
                sheet[f'{date_column_letter}{index + 2}'] = current_date
                sheet[f'{time_column_letter}{index + 2}'] = current_time

                # Esperar un poco para evitar enviar datos demasiado rápido
                # time.sleep(1)
                
                # Enviar el formulario
                if self.is_field_available(driver, "submit_button"):
                    driver.find_element(By.NAME, "submit_button").click()

            # Guardar el archivo Excel con las fechas, horas y números de trámite agregados
            workbook.save(self.excel_file_path)
            
            if self.is_running:
                messagebox.showinfo("Éxito", "Formulario completado y archivo guardado.")

            driver.quit()
        
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.process_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)

    def is_field_available(self, driver, field_name):
        try:
            element = driver.find_element(By.NAME, field_name)
            return element.is_displayed() and element.is_enabled()
        except:
            return False

if __name__ == "__main__":
    root = tk.Tk()
    app = FormFillerApp(root)
    root.mainloop()
